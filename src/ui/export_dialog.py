"""
Sawit Go - TSJ - Export Dialog
Export dialog for generating Excel files
"""

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QPushButton, QComboBox, QLabel, QLineEdit, QFileDialog,
    QDialogButtonBox, QMessageBox, QGroupBox, QRadioButton,
    QCheckBox, QSpinBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from typing import Dict, Any, Optional
from pathlib import Path


class ExportDialog(QDialog):
    """Export dialog for Excel export"""
    
    def __init__(
        self,
        report_type: str = "trial_balance",
        report_data: Optional[Any] = None,
        parent=None
    ):
        super().__init__(parent)
        self.report_type = report_type
        self.report_data = report_data
        
        self.setWindowTitle("Export to Excel")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        self._setup_ui()
    
    def _setup_ui(self):
        """Setup UI components"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        header = QLabel("Export Laporan ke Excel")
        header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(header)
        
        info_group = QGroupBox("Informasi Laporan")
        info_layout = QFormLayout()
        info_layout.setSpacing(8)
        
        self.report_type_label = QLabel(self._get_report_type_name())
        info_layout.addRow("Jenis Laporan:", self.report_type_label)
        
        if self.report_data:
            if hasattr(self.report_data, 'company_name'):
                self.company_label = QLabel(self.report_data.company_name)
            else:
                self.company_label = QLabel("-")
            info_layout.addRow("Perusahaan:", self.company_label)
            
            if hasattr(self.report_data, 'period_name'):
                self.period_label = QLabel(self.report_data.period_name)
            else:
                self.period_label = QLabel("-")
            info_layout.addRow("Period:", self.period_label)
            
            if hasattr(self.report_data, 'items'):
                self.count_label = QLabel(f"{len(self.report_data.items)} item")
            else:
                self.count_label = QLabel("-")
            info_layout.addRow("Jumlah Data:", self.count_label)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        file_group = QGroupBox("Pengaturan File")
        file_layout = QFormLayout()
        file_layout.setSpacing(8)
        
        file_layout.addRow(QLabel("Format:"))
        format_layout = QHBoxLayout()
        
        self.excel_radio = QRadioButton("Microsoft Excel (.xlsx)")
        self.excel_radio.setChecked(True)
        format_layout.addWidget(self.excel_radio)
        format_layout.addStretch()
        file_layout.addRow(format_layout)
        
        file_layout.addRow("Nama File:")
        file_name_layout = QHBoxLayout()
        
        self.file_name_input = QLineEdit()
        self.file_name_input.setText(self._get_default_file_name())
        file_name_layout.addWidget(self.file_name_input)
        
        self.browse_button = QPushButton("Browse...")
        self.browse_button.clicked.connect(self._on_browse)
        file_name_layout.addWidget(self.browse_button)
        
        file_layout.addRow(self.file_name_input)
        file_layout.addRow("", file_name_layout)
        
        self.file_path_label = QLabel("Lokasi: -")
        file_layout.addRow("Lokasi:", self.file_path_label)
        
        self.default_path = str(Path.home() / "Documents" / "SawitGo")
        self.file_path_label.setText(self.default_path)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        options_group = QGroupBox("Opsi Tambahan")
        options_layout = QVBoxLayout()
        options_layout.setSpacing(5)
        
        self.include_header_check = QCheckBox("Sertakan header perusahaan")
        self.include_header_check.setChecked(True)
        options_layout.addWidget(self.include_header_check)
        
        self.include_totals_check = QCheckBox("Sertakan total")
        self.include_totals_check.setChecked(True)
        options_layout.addWidget(self.include_totals_check)
        
        self.open_after_export_check = QCheckBox("Buka file setelah export")
        self.open_after_export_check.setChecked(False)
        options_layout.addWidget(self.open_after_export_check)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self._on_export)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def _get_report_type_name(self) -> str:
        """Get display name for report type"""
        type_names = {
            "trial_balance": "Neraca Saldo (Trial Balance)",
            "journal_book": "Buku Jurnal",
            "ledger": "Buku Besar",
            "journal_list": "Daftar Jurnal"
        }
        return type_names.get(self.report_type, self.report_type)
    
    def _get_default_file_name(self) -> str:
        """Get default file name based on report type"""
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        
        type_names = {
            "trial_balance": f"TrialBalance_{timestamp}",
            "journal_book": f"JournalBook_{timestamp}",
            "ledger": f"Ledger_{timestamp}",
            "journal_list": f"JournalList_{timestamp}"
        }
        
        return type_names.get(self.report_type, f"Report_{timestamp}")
    
    def _on_browse(self):
        """Handle browse button click"""
        file_name, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Simpan File",
            str(Path(self.default_path) / f"{self.file_name_input.text()}.xlsx"),
            "Excel Files (*.xlsx)"
        )
        
        if file_name:
            self.file_path_label.setText(str(Path(file_name).parent))
            self.file_name_input.setText(Path(file_name).stem)
    
    def _on_export(self):
        """Handle export button click"""
        file_name = self.file_name_input.text().strip()
        
        if not file_name:
            QMessageBox.warning(
                self,
                "Validasi",
                "Nama file tidak boleh kosong!"
            )
            return
        
        file_path = str(Path(self.file_path_label.text()) / f"{file_name}.xlsx")
        
        from src.services.export_service import export_service
        
        success = False
        message = ""
        
        if self.report_type == "trial_balance":
            success, message = export_service.export_trial_balance(
                report_data=self.report_data,
                file_path=file_path
            )
        elif self.report_type == "journal_book":
            success, message = export_service.export_journal_book(
                report_data=self.report_data,
                file_path=file_path
            )
        elif self.report_type == "ledger":
            success, message = export_service.export_account_ledger(
                account_data=self.report_data.get('account'),
                ledger_entries=self.report_data.get('entries', []),
                file_path=file_path
            )
        elif self.report_type == "journal_list":
            success, message = export_service.export_journal_entries(
                journal_list=self.report_data.get('journals', []),
                file_path=file_path,
                title=self.report_data.get('title', 'Daftar Jurnal')
            )
        else:
            QMessageBox.critical(
                self,
                "Error",
                f"Tipe laporan tidak dikenali: {self.report_type}"
            )
            return
        
        if success:
            QMessageBox.information(self, "Berhasil", message)
            
            if self.open_after_export_check.isChecked():
                import subprocess
                import platform
                
                try:
                    if platform.system() == "Windows":
                        subprocess.Popen(["start", file_path], shell=True)
                    elif platform.system() == "Darwin":
                        subprocess.Popen(["open", file_path])
                    else:
                        subprocess.Popen(["xdg-open", file_path])
                except Exception:
                    pass
            
            self.accept()
        else:
            QMessageBox.critical(self, "Error", message)
    
    def get_export_path(self) -> str:
        """Get the selected export path"""
        return str(Path(self.file_path_label.text()) / f"{self.file_name_input.text()}.xlsx")


class QuickExportDialog(QDialog):
    """Quick export dialog for common export tasks"""
    
    def __init__(self, company_id: int, parent=None):
        super().__init__(parent)
        self.company_id = company_id
        
        self.setWindowTitle("Quick Export")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        self._setup_ui()
    
    def _setup_ui(self):
        """Setup UI components"""
        layout = QVBoxLayout(self)
        
        header = QLabel("Quick Export")
        header.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(header)
        
        layout.addWidget(QLabel("Pilih jenis laporan untuk di-export:"))
        layout.addSpacing(10)
        
        self.trial_balance_btn = QPushButton("Neraca Saldo (Trial Balance)")
        self.trial_balance_btn.clicked.connect(lambda: self._quick_export("trial_balance"))
        layout.addWidget(self.trial_balance_btn)
        
        self.journal_book_btn = QPushButton("Buku Jurnal")
        self.journal_book_btn.clicked.connect(lambda: self._quick_export("journal_book"))
        layout.addWidget(self.journal_book_btn)
        
        self.journal_list_btn = QPushButton("Daftar Jurnal")
        self.journal_list_btn.clicked.connect(lambda: self._quick_export("journal_list"))
        layout.addWidget(self.journal_list_btn)
        
        self.gl_accounts_btn = QPushButton("Daftar Akun GL")
        self.gl_accounts_btn.clicked.connect(lambda: self._quick_export("gl_accounts"))
        layout.addWidget(self.gl_accounts_btn)
        
        layout.addStretch()
        
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def _quick_export(self, report_type: str):
        """Handle quick export selection"""
        from datetime import datetime
        from src.services import export_service
        
        default_dir = str(Path.home() / "Documents" / "SawitGo")
        Path(default_dir).mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan File",
            str(Path(default_dir) / f"{report_type}_{timestamp}.xlsx"),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        success = False
        message = ""
        
        if report_type == "trial_balance":
            from src.services.report_service import report_service
            from src.database.session import DatabaseSession
            from src.models import Period
            
            session = DatabaseSession.get_session()
            try:
                period = session.query(Period).filter(
                    Period.company_id == self.company_id
                ).first()
                
                if period:
                    report_data = report_service.generate_trial_balance(
                        company_id=self.company_id,
                        period_id=period.id
                    )
                    
                    if report_data:
                        success, message = export_service.export_trial_balance(
                            report_data=report_data,
                            file_path=file_path
                        )
            finally:
                session.close()
        
        elif report_type == "journal_list":
            from src.services.journal_service import journal_service
            
            journals = journal_service.get_journals(company_id=self.company_id)
            success, message = export_service.export_journal_entries(
                journal_list=journals,
                file_path=file_path
            )
        
        elif report_type == "gl_accounts":
            from src.services.gl_account_service import gl_account_service
            
            accounts = gl_account_service.get_all(self.company_id)
            success, message = self._export_gl_accounts(accounts, file_path)
        
        if success:
            QMessageBox.information(self, "Berhasil", message)
        else:
            QMessageBox.critical(self, "Error", message)
    
    def _export_gl_accounts(self, accounts: list, file_path: str) -> tuple:
        """Export GL accounts to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Chart of Accounts"
            
            headers = ["Kode", "Nama Akun", "Tipe", "Saldo Normal", "Aktif"]
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            for row, account in enumerate(accounts, 2):
                ws.cell(row=row, column=1, value=account.get('code', ''))
                ws.cell(row=row, column=2, value=account.get('name', ''))
                ws.cell(row=row, column=3, value=account.get('account_type', ''))
                ws.cell(row=row, column=4, value=account.get('normal_balance', ''))
                ws.cell(row=row, column=5, value="Ya" if account.get('is_active') else "Tidak")
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 8
            
            wb.save(file_path)
            return True, f"File berhasil disimpan ke {file_path}"
            
        except Exception as e:
            return False, f"Error: {str(e)}"


class ExportProgressDialog(QDialog):
    """Progress dialog for export operations"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Exporting...")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout(self)
        
        self.status_label = QLabel("Memproses...")
        layout.addWidget(self.status_label)
        
        self.progress_label = QLabel("0%")
        layout.addWidget(self.progress_label)
    
    def update_progress(self, value: int, status: str = ""):
        """Update progress display"""
        self.progress_label.setText(f"{value}%")
        if status:
            self.status_label.setText(status)
        
        from PyQt6.QtCore import QCoreApplication
        QCoreApplication.processEvents()
    
    def show_success(self, message: str):
        """Show success message"""
        self.status_label.setText(message)
        self.progress_label.setText("100%")
        QMessageBox.information(self, "Berhasil", message)
    
    def show_error(self, message: str):
        """Show error message"""
        QMessageBox.critical(self, "Error", message)
