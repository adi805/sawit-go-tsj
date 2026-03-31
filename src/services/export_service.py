"""
Sawit Go - TSJ - Export Service
Excel export functionality using openpyxl
"""

from typing import List, Dict, Any, Optional, Tuple
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

logger = logging.getLogger(__name__)


class ExportService:
    """Handle Excel export operations"""

    def __init__(self):
        self.default_font = Font(name="Arial", size=10)
        self.header_font = Font(name="Arial", size=11, bold=True)
        self.title_font = Font(name="Arial", size=14, bold=True)
        self.bold_font = Font(name="Arial", size=10, bold=True)
        
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.header_font_white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        self.total_fill = PatternFill(
            start_color="D9E2F3",
            end_color="D9E2F3",
            fill_type="solid"
        )
        
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")

    def _format_currency(self, value: float) -> str:
        """Format number as currency"""
        if value is None:
            return ""
        return f"Rp {value:,.2f}"

    def _format_number(self, value: float, decimal_places: int = 2) -> str:
        """Format number with decimal places"""
        if value is None:
            return ""
        return f"{value:,.{decimal_places}f}"

    def export_trial_balance(
        self,
        report_data: Any,
        file_path: str
    ) -> Tuple[bool, str]:
        """
        Export trial balance report to Excel.
        
        Args:
            report_data: TrialBalanceReport object
            file_path: Output file path
            
        Returns:
            Tuple of (success, message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            
            row = 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value="NERACA SALDO")
            cell.font = self.title_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=report_data.company_name)
            cell.font = self.header_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f"Periode: {report_data.period_name}")
            cell.font = self.default_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f"Tanggal: {report_data.start_date} s/d {report_data.end_date}")
            cell.font = self.default_font
            cell.alignment = self.center_align
            row += 2
            
            headers = ["No", "Kode Akun", "Nama Akun", "Debit", "Kredit", "Saldo"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font_white
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_align
            row += 1
            
            no = 1
            for item in report_data.items:
                ws.cell(row=row, column=1, value=no).alignment = self.center_align
                ws.cell(row=row, column=2, value=item["account_code"]).alignment = self.left_align
                ws.cell(row=row, column=3, value=item["account_name"]).alignment = self.left_align
                
                debit_cell = ws.cell(row=row, column=4, value=item["total_debit"])
                debit_cell.number_format = '#,##0.00'
                debit_cell.alignment = self.right_align
                
                credit_cell = ws.cell(row=row, column=5, value=item["total_credit"])
                credit_cell.number_format = '#,##0.00'
                credit_cell.alignment = self.right_align
                
                balance_cell = ws.cell(row=row, column=6, value=item["ending_balance"])
                balance_cell.number_format = '#,##0.00'
                balance_cell.alignment = self.right_align
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                no += 1
                row += 1
            
            row += 1
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = self.total_fill
                cell.border = self.thin_border
                if col == 1:
                    cell.value = ""
                elif col == 2:
                    cell.value = "TOTAL"
                    cell.font = self.bold_font
                elif col == 4:
                    cell.value = report_data.total_debit
                    cell.number_format = '#,##0.00'
                    cell.font = self.bold_font
                    cell.alignment = self.right_align
                elif col == 5:
                    cell.value = report_data.total_credit
                    cell.number_format = '#,##0.00'
                    cell.font = self.bold_font
                    cell.alignment = self.right_align
                elif col == 6:
                    cell.value = ""
            
            row += 2
            ws.cell(row=row, column=1, value=f"Dicetak: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            
            wb.save(file_path)
            
            return True, f"File berhasil disimpan ke {file_path}"
            
        except Exception as e:
            logger.error(f"Export trial balance error: {e}")
            return False, f"Error: {str(e)}"

    def export_journal_book(
        self,
        report_data: Any,
        file_path: str
    ) -> Tuple[bool, str]:
        """
        Export journal book report to Excel.
        
        Args:
            report_data: JournalBookReport object
            file_path: Output file path
            
        Returns:
            Tuple of (success, message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Buku Jurnal"
            
            row = 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value="BUKU JURNAL UMUM")
            cell.font = self.title_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=report_data.company_name)
            cell.font = self.header_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f"Periode: {report_data.period_name}")
            cell.font = self.default_font
            cell.alignment = self.center_align
            row += 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=f"Tanggal: {report_data.start_date} s/d {report_data.end_date}")
            cell.font = self.default_font
            cell.alignment = self.center_align
            row += 2
            
            for entry in report_data.entries:
                ws.cell(row=row, column=1, value=f"No Jurnal: {entry['journal_no']}")
                ws.cell(row=row, column=1).font = self.bold_font
                row += 1
                
                ws.cell(row=row, column=1, value=f"Tanggal: {entry['date']}")
                ws.cell(row=row, column=3, value=f"Referensi: {entry['reference'] or '-'}")
                row += 1
                
                ws.cell(row=row, column=1, value=f"Keterangan: {entry['description']}")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 1
                
                headers = ["No", "Kode", "Akun", "Keterangan", "Debit", "Kredit"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.header_font_white
                    cell.fill = self.header_fill
                    cell.border = self.thin_border
                    cell.alignment = self.center_align
                row += 1
                
                line_no = 1
                for line in entry["lines"]:
                    ws.cell(row=row, column=1, value=line_no).alignment = self.center_align
                    ws.cell(row=row, column=2, value=line["account_code"]).alignment = self.left_align
                    ws.cell(row=row, column=3, value=line["account_name"]).alignment = self.left_align
                    ws.cell(row=row, column=4, value=line["description"]).alignment = self.left_align
                    
                    if line["debit"] > 0:
                        debit_cell = ws.cell(row=row, column=5, value=line["debit"])
                        debit_cell.number_format = '#,##0.00'
                        debit_cell.alignment = self.right_align
                    else:
                        ws.cell(row=row, column=5, value="")
                    
                    if line["credit"] > 0:
                        credit_cell = ws.cell(row=row, column=6, value=line["credit"])
                        credit_cell.number_format = '#,##0.00'
                        credit_cell.alignment = self.right_align
                    else:
                        ws.cell(row=row, column=6, value="")
                    
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).border = self.thin_border
                    
                    line_no += 1
                    row += 1
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.total_fill
                    cell.border = self.thin_border
                    if col == 1:
                        cell.value = ""
                    elif col == 4:
                        cell.value = "TOTAL"
                        cell.font = self.bold_font
                    elif col == 5:
                        cell.value = entry["total_debit"]
                        cell.number_format = '#,##0.00'
                        cell.font = self.bold_font
                        cell.alignment = self.right_align
                    elif col == 6:
                        cell.value = entry["total_credit"]
                        cell.number_format = '#,##0.00'
                        cell.font = self.bold_font
                        cell.alignment = self.right_align
                
                row += 2
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            
            wb.save(file_path)
            
            return True, f"File berhasil disimpan ke {file_path}"
            
        except Exception as e:
            logger.error(f"Export journal book error: {e}")
            return False, f"Error: {str(e)}"

    def export_account_ledger(
        self,
        account_data: Dict[str, Any],
        ledger_entries: List[Dict[str, Any]],
        file_path: str
    ) -> Tuple[bool, str]:
        """
        Export account ledger to Excel.
        
        Args:
            account_data: Account information dict
            ledger_entries: List of ledger entry dicts
            file_path: Output file path
            
        Returns:
            Tuple of (success, message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Ledger {account_data['code']}"
            
            row = 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value="BUKU BESAR")
            cell.font = self.title_font
            cell.alignment = self.center_align
            row += 2
            
            ws.cell(row=row, column=1, value="Akun:")
            ws.cell(row=row, column=2, value=f"{account_data['code']} - {account_data['name']}")
            ws.cell(row=row, column=2).font = self.bold_font
            row += 1
            
            ws.cell(row=row, column=1, value="Tipe:")
            ws.cell(row=row, column=2, value=account_data.get('account_type', 'DETAIL'))
            row += 1
            
            ws.cell(row=row, column=1, value="Saldo Normal:")
            ws.cell(row=row, column=2, value=account_data.get('normal_balance', 'DEBIT'))
            row += 2
            
            headers = ["Tanggal", "No Jurnal", "Keterangan", "Referensi", "Debit", "Kredit", "Saldo"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font_white
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_align
            row += 1
            
            running_balance = account_data.get('initial_balance', 0)
            
            for entry in ledger_entries:
                ws.cell(row=row, column=1, value=entry.get('date', '')).alignment = self.center_align
                ws.cell(row=row, column=2, value=entry.get('journal_no', '')).alignment = self.center_align
                ws.cell(row=row, column=3, value=entry.get('description', '')).alignment = self.left_align
                ws.cell(row=row, column=4, value=entry.get('reference', '')).alignment = self.left_align
                
                debit = entry.get('debit', 0)
                credit = entry.get('credit', 0)
                
                if debit > 0:
                    debit_cell = ws.cell(row=row, column=5, value=debit)
                    debit_cell.number_format = '#,##0.00'
                    debit_cell.alignment = self.right_align
                    running_balance += debit
                else:
                    ws.cell(row=row, column=5, value="")
                
                if credit > 0:
                    credit_cell = ws.cell(row=row, column=6, value=credit)
                    credit_cell.number_format = '#,##0.00'
                    credit_cell.alignment = self.right_align
                    running_balance -= credit
                else:
                    ws.cell(row=row, column=6, value="")
                
                balance_cell = ws.cell(row=row, column=7, value=running_balance)
                balance_cell.number_format = '#,##0.00'
                balance_cell.alignment = self.right_align
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                row += 1
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18
            
            row += 1
            ws.cell(row=row, column=1, value=f"Dicetak: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            wb.save(file_path)
            
            return True, f"File berhasil disimpan ke {file_path}"
            
        except Exception as e:
            logger.error(f"Export ledger error: {e}")
            return False, f"Error: {str(e)}"

    def export_journal_entries(
        self,
        journal_list: List[Dict[str, Any]],
        file_path: str,
        title: str = "Daftar Jurnal"
    ) -> Tuple[bool, str]:
        """
        Export list of journal entries to Excel.
        
        Args:
            journal_list: List of journal entry dicts
            file_path: Output file path
            title: Report title
            
        Returns:
            Tuple of (success, message)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Journals"
            
            row = 1
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = self.title_font
            cell.alignment = self.center_align
            row += 2
            
            headers = ["No", "No Jurnal", "Tanggal", "Referensi", "Keterangan", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font_white
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_align
            row += 1
            
            no = 1
            for journal in journal_list:
                ws.cell(row=row, column=1, value=no).alignment = self.center_align
                ws.cell(row=row, column=2, value=journal.get('journal_no', '')).alignment = self.center_align
                ws.cell(row=row, column=3, value=journal.get('date', '')).alignment = self.center_align
                ws.cell(row=row, column=4, value=journal.get('reference', '')).alignment = self.left_align
                ws.cell(row=row, column=5, value=journal.get('description', '')).alignment = self.left_align
                ws.cell(row=row, column=6, value=journal.get('status', '')).alignment = self.center_align
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                no += 1
                row += 1
            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 10
            
            row += 1
            ws.cell(row=row, column=1, value=f"Total: {len(journal_list)} jurnal")
            ws.cell(row=row, column=1).font = self.bold_font
            
            wb.save(file_path)
            
            return True, f"File berhasil disimpan ke {file_path}"
            
        except Exception as e:
            logger.error(f"Export journal list error: {e}")
            return False, f"Error: {str(e)}"


export_service = ExportService()
